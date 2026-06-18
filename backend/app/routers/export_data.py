import io
from datetime import datetime
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session, joinedload

from ..auth import require_teacher
from ..database import get_db
from ..models import Exam, ExamSession, SessionStatus, TypingRecord, User, UserRole
from ..permissions import apply_student_scope, ensure_class_access, filter_class_name, get_assigned_classes

router = APIRouter(prefix="/api/export", tags=["数据导出"])


def _workbook_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ascii_name = "export.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={ascii_name}; filename*=UTF-8''{encoded}",
        },
    )


def _sheet_title(wb: Workbook, title: str):
    sheet = wb.active
    sheet.title = title[:31]
    return sheet


@router.get("/students")
def export_students(
    class_name: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_teacher),
):
    class_name = filter_class_name(current_user, class_name)
    query = db.query(User)
    query = apply_student_scope(query, current_user, class_name)
    students = query.order_by(User.class_name, User.id).all()

    wb = Workbook()
    ws = _sheet_title(wb, "学生名单")
    ws.append(["用户名", "姓名", "班级", "状态", "创建时间"])
    for s in students:
        ws.append([
            s.username,
            s.name,
            s.class_name,
            "正常" if s.is_active else "禁用",
            s.created_at.strftime("%Y-%m-%d %H:%M") if s.created_at else "",
        ])
    suffix = class_name or "全部"
    return _workbook_response(wb, f"学生名单_{suffix}.xlsx")


@router.get("/exams/{exam_id}/results")
def export_exam_results(
    exam_id: int,
    class_name: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_teacher),
):
    class_name = filter_class_name(current_user, class_name)
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    sessions = (
        db.query(ExamSession)
        .options(joinedload(ExamSession.student))
        .filter(ExamSession.exam_id == exam_id)
        .all()
    )
    classes = get_assigned_classes(current_user)
    rows = []
    for s in sessions:
        stu = s.student
        if not stu:
            continue
        if classes is not None and stu.class_name not in classes:
            continue
        if class_name and stu.class_name != class_name:
            continue
        rows.append(s)

    wb = Workbook()
    ws = _sheet_title(wb, "考试成绩")
    ws.append(["考试名称", exam.title])
    ws.append(["姓名", "班级", "状态", "得分", "满分", "及格", "开始时间", "提交时间"])
    for s in rows:
        stu = s.student
        passed = s.total_score >= exam.pass_score if s.status == SessionStatus.submitted else False
        ws.append([
            stu.name,
            stu.class_name,
            "已提交" if s.status == SessionStatus.submitted else "答题中",
            s.total_score if s.status == SessionStatus.submitted else "",
            s.max_score,
            "是" if passed else ("否" if s.status == SessionStatus.submitted else ""),
            s.start_time.strftime("%Y-%m-%d %H:%M") if s.start_time else "",
            s.submit_time.strftime("%Y-%m-%d %H:%M") if s.submit_time else "",
        ])
    return _workbook_response(wb, f"考试成绩_{exam.title}.xlsx")


@router.get("/typing/records")
def export_typing_records(
    class_name: str | None = None,
    student_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_teacher),
):
    class_name = filter_class_name(current_user, class_name)
    query = (
        db.query(TypingRecord)
        .options(joinedload(TypingRecord.student), joinedload(TypingRecord.text))
        .order_by(TypingRecord.id.desc())
    )
    if student_id:
        student = db.query(User).filter(User.id == student_id, User.role == UserRole.student).first()
        if not student:
            raise HTTPException(status_code=404, detail="学生不存在")
        ensure_class_access(current_user, student.class_name)
        query = query.filter(TypingRecord.student_id == student_id)
    else:
        classes = get_assigned_classes(current_user)
        if classes is not None:
            if class_name:
                query = query.join(User).filter(User.class_name == class_name)
            else:
                query = query.join(User).filter(User.class_name.in_(classes))
        elif class_name:
            query = query.join(User).filter(User.class_name == class_name)

    records = query.limit(2000).all()
    wb = Workbook()
    ws = _sheet_title(wb, "打字记录")
    ws.append(["姓名", "班级", "文章", "模式", "评分", "速度", "准确率", "等级", "时间"])
    source_map = {"practice": "自由练习", "test": "5分钟测试", "class_test": "课堂测试"}
    for r in records:
        stu = r.student
        ws.append([
            stu.name if stu else "",
            stu.class_name if stu else "",
            r.text.title if r.text else "自定义",
            source_map.get(r.source, r.source),
            r.score,
            r.wpm,
            r.accuracy,
            r.level,
            r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
        ])
    suffix = class_name or "全部"
    return _workbook_response(wb, f"打字记录_{suffix}.xlsx")
